import requests
import json
import jsonpath
import openpyxl

def test_add_multiple_students():
    # API
    api_url= "http://thetestingworldapi.com/api/studentsDetails"
    f = open('C:/Users/TestingWorld/Desktop/TASK_API/AddNewStudent.json')
    json_request = json.loads(f.read())

    # Excel Code
    wk = openpyxl.load_workbook('C:/Users/TestingWorld/Desktop/TASK_API/TestData.xlsx')
    sh = wk['Sheet1']
    rows = sh.max_row


    for i in range(2,rows+1):
         cell_first_name = sh.cell(row=i, column=1)
         cell_mid_name = sh.cell(row=i, column=2)
         cell_last_name = sh.cell(row=i, column=3)
         cell_dob = sh.cell(row=i, column=4)

         json_request['first_name']=cell_first_name.value
         json_request['middle_name']=cell_mid_name.value
         json_request['last_name']=cell_last_name.value
         json_request['date_of_birth']=cell_dob.value

         response = requests.post(api_url, json_request)

         print(response.text)
         print(response.status_code)
         assert response.status_code == 201
